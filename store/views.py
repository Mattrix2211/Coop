from decimal import Decimal
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import F, Sum
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.conf import settings
import os

from django.contrib.auth.models import User, Group
from .decorators import admin_required

DENOMINATIONS = [
	Decimal('50'), Decimal('20'), Decimal('10'), Decimal('5'),
	Decimal('2'), Decimal('1'), Decimal('0.50'), Decimal('0.20'), Decimal('0.10'), Decimal('0.05')
]

from .forms import (
	ProductForm,
	CustomerForm,
	SaleForm,
	SaleItemFormSet,
)
from .models import Product, Customer, Sale, Payment, Category, record_payment_movement, CashSession, CashMovement, get_active_cash_session, StockMovement


@login_required
def home(request):
	return redirect("sale_create")


# Produits
@login_required
def product_list(request):
	q = request.GET.get("q", "").strip()
	category_id = request.GET.get("category")
	products = Product.objects.select_related("category").order_by("category__name", "name")
	if q:
		products = products.filter(name__icontains=q)
	if category_id:
		products = products.filter(category_id=category_id)
	categories = Category.objects.all()
	return render(request, "store/product_list.html", {"products": products, "q": q, "category_id": category_id, "categories": categories})

@admin_required
def product_create(request):
	if request.method == "POST":
		data = request.POST.copy()
		files = request.FILES.copy()
		image_data = data.get("image_data")
		if image_data and image_data.startswith("data:image/"):
			try:
				import base64
				header, b64data = image_data.split(",", 1)
				ext = "png"
				if "jpeg" in header or "jpg" in header:
					ext = "jpg"
				files["image"] = ContentFile(base64.b64decode(b64data), name=f"image_recadree.{ext}")
			except Exception:
				pass
		form = ProductForm(data, files)
		if form.is_valid():
			form.save()
			messages.success(request, "Produit ajouté.")
			return redirect("product_list")
	else:
		form = ProductForm()
	return render(request, "store/product_form.html", {"form": form, "title": "Nouveau produit"})

@admin_required
def product_update(request, pk):
	product = get_object_or_404(Product, pk=pk)
	if request.method == "POST":
		form = ProductForm(request.POST, request.FILES, instance=product)
		if form.is_valid():
			form.save()
			messages.success(request, "Produit mis à jour.")
			return redirect("product_list")
	else:
		form = ProductForm(instance=product)
	return render(request, "store/product_form.html", {"form": form, "title": f"Modifier {product.name}"})

@admin_required
def product_delete(request, pk):
	product = get_object_or_404(Product, pk=pk)
	if request.method == "POST":
		product.delete()
		messages.success(request, "Produit supprimé.")
		return redirect("product_list")
	return render(request, "store/confirm_delete.html", {"object": product, "type": "produit", "cancel_url": reverse("product_list")})

@login_required
@transaction.atomic
def product_add_stock(request, pk):
	product = get_object_or_404(Product, pk=pk)
	if request.method == "POST":
		try:
			add_qty = int(request.POST.get("add_quantity", "0") or "0")
		except Exception:
			add_qty = 0
		if add_qty <= 0:
			messages.error(request, "Quantité invalide.")
		else:
			Product.objects.filter(pk=product.pk).update(quantity=F('quantity') + add_qty)
			StockMovement.objects.create(product=product, change=add_qty, reason=StockMovement.RESTOCK, note="Réassort")
			messages.success(request, f"+{add_qty} au stock de « {product.name} ».")
	return redirect("product_list")

@admin_required
@transaction.atomic
def product_bulk_move(request):
	if request.method == "POST":
		ids = request.POST.getlist("ids")
		category_id = request.POST.get("target_category")
		if not ids or not category_id:
			messages.error(request, "Sélection et catégorie requises.")
			return redirect("product_list")
		Product.objects.filter(id__in=ids).update(category_id=category_id)
		messages.success(request, "Produits déplacés.")
	return redirect("product_list")

# Clients
@login_required
def customer_list(request):
	customers = Customer.objects.all()
	return render(request, "store/customer_list.html", {"customers": customers})

@login_required
def customer_detail(request, pk):
	customer = get_object_or_404(Customer, pk=pk)
	sales = list(
		Sale.objects.filter(customer=customer)
		.prefetch_related('payments', 'items__product')
		.order_by('-created_at')
	)
	total_purchased = sum(s.total for s in sales)
	total_paid = sum(s.amount_paid for s in sales)
	total_debt = sum(s.balance for s in sales)
	return render(request, 'store/customer_detail.html', {
		'customer': customer,
		'sales': sales,
		'total_purchased': total_purchased,
		'total_paid': total_paid,
		'total_debt': total_debt,
		'methods': Payment.METHOD_CHOICES,
	})

@login_required
def customer_create(request):
	if request.method == "POST":
		form = CustomerForm(request.POST)
		if form.is_valid():
			form.save()
			messages.success(request, "Client ajouté.")
			return redirect("customer_list")
	else:
		form = CustomerForm()
	return render(request, "store/customer_form.html", {"form": form, "title": "Nouveau client"})

@login_required
def customer_update(request, pk):
	customer = get_object_or_404(Customer, pk=pk)
	if request.method == "POST":
		form = CustomerForm(request.POST, instance=customer)
		if form.is_valid():
			form.save()
			messages.success(request, "Client mis à jour.")
			return redirect("customer_list")
	else:
		form = CustomerForm(instance=customer)
	return render(request, "store/customer_form.html", {"form": form, "title": f"Modifier {customer}"})

@admin_required
@require_POST
def customer_delete(request, pk):
	customer = get_object_or_404(Customer, pk=pk)
	name = str(customer)
	customer.delete()
	messages.success(request, f"Client « {name} » supprimé.")
	return redirect("customer_list")

@admin_required
@require_POST
def customer_delete_all(request):
	count, _ = Customer.objects.all().delete()
	messages.success(request, f"{count} client(s) supprimé(s).")
	return redirect("customer_list")

def _build_denom_entries():
	"""Construit la liste des dénominations avec l'image disponible (.png prioritaire puis .svg)."""
	entries = []
	for d in DENOMINATIONS:
		val_str = str(d).replace('.', '_')
		png_path = settings.BASE_DIR / 'static' / 'img' / 'denoms' / f'{val_str}.png'
		svg_path = settings.BASE_DIR / 'static' / 'img' / 'denoms' / f'{val_str}.svg'
		if png_path.exists():
			img_rel = f'img/denoms/{val_str}.png'
		elif svg_path.exists():
			img_rel = f'img/denoms/{val_str}.svg'
		else:
			img_rel = None
		entries.append({
			'value': d,
			'safe': val_str,
			'type': 'bill' if d >= Decimal('5') else 'coin',
			'img': img_rel,
		})
	return entries

@login_required
@transaction.atomic
def sale_confirm(request):
	"""Crée la vente et enregistre immédiatement le paiement (s’il y en a un)."""
	if request.method != 'POST':
		return JsonResponse({'ok': False, 'error': 'Méthode non autorisée.'}, status=405)
	form = SaleForm(request.POST)
	formset = SaleItemFormSet(request.POST, prefix='items')
	if not (form.is_valid() and formset.is_valid()):
		return JsonResponse({'ok': False, 'error': 'Formulaire invalide.'}, status=400)
	stock_updates = []
	items_data = []
	total = Decimal('0')
	for f in formset:
		if not f.cleaned_data or f.cleaned_data.get('DELETE'):
			continue
		product = f.cleaned_data['product']
		qty = f.cleaned_data['quantity']
		if qty <= 0:
			return JsonResponse({'ok': False, 'error': 'Quantité invalide.'}, status=400)
		if product.quantity < qty:
			return JsonResponse({'ok': False, 'error': f'Stock insuffisant pour {product.name}.'}, status=400)
		stock_updates.append((product, product.quantity - qty))
		items_data.append({'product': product, 'quantity': qty, 'unit_price': product.price})
		total += product.price * qty
	if not items_data:
		return JsonResponse({'ok': False, 'error': 'Aucun article.'}, status=400)
	if not form.cleaned_data.get('customer'):
		return JsonResponse({'ok': False, 'error': 'Client requis.'}, status=400)
	sale: Sale = form.save()
	for data in items_data:
		sale.items.create(product=data['product'], quantity=data['quantity'], unit_price=data['unit_price'])
		StockMovement.objects.create(product=data['product'], change=-int(data['quantity']), reason=StockMovement.SALE, sale=sale, note=f'Vente #{sale.id}')
	for product, new_qty in stock_updates:
		Product.objects.filter(pk=product.pk).update(quantity=new_qty)
	method = request.POST.get('method') or Payment.CASH
	tendered_breakdown = {}
	amount_tendered = Decimal('0')
	if method == Payment.CASH:
		for denom in DENOMINATIONS:
			fname = f'denom_{str(denom).replace(".", "_")}'
			try:
				qty = int(request.POST.get(fname, '0') or '0')
			except Exception:
				qty = 0
			if qty > 0:
				tendered_breakdown[str(denom)] = qty
		amount_tendered = sum((Decimal(k) * v for k, v in tendered_breakdown.items()), start=Decimal('0'))
		amount_paid = total if amount_tendered >= total else amount_tendered
	else:
		try:
			amount_paid = Decimal(request.POST.get('amount', '0') or '0')
		except Exception:
			amount_paid = Decimal('0')
	if amount_paid > 0:
		payment = Payment.objects.create(sale=sale, amount=amount_paid, method=method)
		change_breakdown = {}
		change_amount = Decimal('0')
		if method == Payment.CASH and amount_tendered > total:
			change_amount = amount_tendered - total
			manual_breakdown = {}
			for d in DENOMINATIONS:
				cf = f'change_denom_{str(d).replace(".", "_")}'
				try:
					q = int(request.POST.get(cf, '0') or '0')
				except Exception:
					q = 0
				if q>0:
					manual_breakdown[str(d)] = q
			if manual_breakdown:
				manual_total = sum((Decimal(k)*v for k,v in manual_breakdown.items()), start=Decimal('0'))
				if abs(manual_total - change_amount) > Decimal('0.01'):
					return JsonResponse({'ok': False, 'error': 'Rendu incohérent (total différent).'}, status=400)
				change_breakdown = manual_breakdown
			else:
				session = get_active_cash_session()
				available = {}
				if session and session.current_counts:
					available = {k: int(v) for k, v in (session.current_counts.get('cash') or {}).items()}
				for k, v in tendered_breakdown.items():
					available[k] = available.get(k, 0) + v
				remaining = change_amount
				for denom in sorted(DENOMINATIONS, reverse=True):
					if remaining <= 0:
						break
					denom_str = str(denom)
					units_available = available.get(denom_str, 0)
					needed = int(remaining // denom)
					use = min(needed, units_available)
					if use > 0:
						change_breakdown[denom_str] = use
						remaining -= denom * use
		record_payment_movement(payment, tendered_breakdown=tendered_breakdown if method == Payment.CASH else None, change_breakdown=change_breakdown if change_amount > 0 else None, change_amount=change_amount if change_amount > 0 else None)
	# Construire reçu
	cash_added_by_payment = {}
	cash_change_breakdowns = []
	for m in CashMovement.objects.filter(sale=sale).select_related('payment'):
		if m.movement_type == CashMovement.PAYMENT and m.method == Payment.CASH and m.counts_snapshot and 'cash_added' in m.counts_snapshot:
			cash_added_by_payment[m.payment_id] = m.counts_snapshot['cash_added']
		elif m.movement_type == CashMovement.CHANGE and m.counts_snapshot and 'cash_change' in m.counts_snapshot:
			cash_change_breakdowns.append(m.counts_snapshot['cash_change'])
	payment_entries = []
	for p in sale.payments.all():
		payment_entries.append({'payment': p, 'cash_added': cash_added_by_payment.get(p.id)})
	receipt_html = render(request, 'store/_sale_receipt.html', {
		'sale': sale,
		'payment_entries': payment_entries,
		'cash_change_breakdowns': cash_change_breakdowns,
	}).content.decode('utf-8')
	return JsonResponse({'ok': True, 'sale_id': sale.id, 'receipt_html': receipt_html})


# Ventes
@login_required
def sale_create(request):
	form = SaleForm()
	formset = SaleItemFormSet(prefix='items')
	gallery_products = Product.objects.order_by("category__name", "name").select_related("category")
	denom_entries = _build_denom_entries()
	session = get_active_cash_session()
	cash_current_counts = {}
	if session and session.current_counts:
		cash_current_counts = session.current_counts.get('cash') or {}
	return render(request, 'store/sale_form.html', {
		'form': form,
		'formset': formset,
		'gallery_products': gallery_products,
		'denom_entries': denom_entries,
		'cash_current_counts': cash_current_counts,
	})


@login_required
def sale_list(request):
	q_customer = request.GET.get('customer', '').strip()
	q_status = request.GET.get('status', '')

	qs = (
		Sale.objects.select_related("customer")
		.prefetch_related('payments', 'items')
		.order_by("-created_at")
	)
	if q_customer:
		qs = qs.filter(
			Q(customer__first_name__icontains=q_customer) |
			Q(customer__last_name__icontains=q_customer)
		)

	sales_all = list(qs)
	if q_status == 'paid':
		sales_all = [s for s in sales_all if s.balance <= 0]
	elif q_status == 'unpaid':
		sales_all = [s for s in sales_all if s.balance > 0]

	paginator = Paginator(sales_all, 25)
	page = paginator.get_page(request.GET.get('page', 1))

	ctx = {
		"sales": page,
		"methods": Payment.METHOD_CHOICES,
		"q_customer": q_customer,
		"q_status": q_status,
	}
	return render(request, "store/sale_list.html", ctx)


@admin_required
@transaction.atomic
@require_POST
def clear_sales_history(request):
	"""Supprime toutes les ventes (historique), incluant lignes et paiements.

	Attention: cela ne remet pas à niveau les stocks produits.
	"""
	Sale.objects.all().delete()
	messages.success(request, "Historique des ventes supprimé.")
	return redirect("sale_list")


# Vue sale_detail supprimée: les détails s'affichent via un reçu en modale


@login_required
def sale_receipt_fragment(request, pk):
	"""Retourne un fragment HTML (reçu) pour affichage dans une modale (inclut ventilations espèces)."""
	sale = get_object_or_404(Sale.objects.select_related('customer'), pk=pk)
	cash_added_by_payment = {}
	cash_change_breakdowns = []
	for m in CashMovement.objects.filter(sale=sale).select_related('payment'):
		if m.movement_type == CashMovement.PAYMENT and m.method == Payment.CASH and m.counts_snapshot and 'cash_added' in m.counts_snapshot:
			cash_added_by_payment[m.payment_id] = m.counts_snapshot['cash_added']
		elif m.movement_type == CashMovement.CHANGE and m.counts_snapshot and 'cash_change' in m.counts_snapshot:
			cash_change_breakdowns.append(m.counts_snapshot['cash_change'])
	return render(request, 'store/_sale_receipt.html', {
		'sale': sale,
		'cash_added_by_payment': cash_added_by_payment,
		'cash_change_breakdowns': cash_change_breakdowns,
	})

@login_required
@transaction.atomic
def pay_sale(request, pk):
	sale = get_object_or_404(Sale, pk=pk)
	if request.method != 'POST':
		return JsonResponse({'ok': False, 'error': 'Méthode non autorisée.'}, status=405)
	method = request.POST.get('method') or Payment.CASH
	balance = sale.balance
	if balance <= 0:
		return JsonResponse({'ok': False, 'error': 'Vente déjà réglée.'}, status=400)
	tendered_breakdown = {}
	amount_tendered = Decimal('0')
	if method == Payment.CASH:
		for denom in DENOMINATIONS:
			fname = f'denom_{str(denom).replace(".", "_")}'
			try:
				qty = int(request.POST.get(fname, '0') or '0')
			except Exception:
				qty = 0
			if qty > 0:
				key = str(denom)
				tendered_breakdown[key] = qty
		amount_tendered = sum((Decimal(k) * v for k, v in tendered_breakdown.items()), start=Decimal('0'))
		if amount_tendered <= 0:
			return JsonResponse({'ok': False, 'error': 'Montant espèces invalide.'}, status=400)
		amount_paid = balance if amount_tendered >= balance else amount_tendered
	else:
		try:
			amount_paid = Decimal(request.POST.get('amount', '0') or '0')
		except Exception:
			amount_paid = Decimal('0')
		if amount_paid <= 0:
			return JsonResponse({'ok': False, 'error': 'Montant paiement invalide.'}, status=400)
	payment = Payment.objects.create(sale=sale, amount=amount_paid, method=method)
	change_breakdown = {}
	change_amount = Decimal('0')
	if method == Payment.CASH and amount_tendered > balance:
		change_amount = amount_tendered - balance
		session = get_active_cash_session()
		available = {}
		if session and session.current_counts:
			available = {k: int(v) for k, v in (session.current_counts.get('cash') or {}).items()}
		for k, v in tendered_breakdown.items():
			available[k] = available.get(k, 0) + int(v)
		remaining = change_amount
		for denom in sorted(DENOMINATIONS, reverse=True):
			denom_str = str(denom)
			if remaining <= 0:
				break
			units_available = available.get(denom_str, 0)
			needed = int(remaining // denom)
			use = min(needed, units_available)
			if use > 0:
				change_breakdown[denom_str] = use
				remaining -= denom * use
	record_payment_movement(payment, tendered_breakdown=tendered_breakdown if method == Payment.CASH else None, change_breakdown=change_breakdown if change_amount > 0 else None, change_amount=change_amount if change_amount > 0 else None)
	receipt_html = render(request, 'store/_sale_receipt.html', {
		'sale': sale,
		'cash_added_by_payment': {payment.id: tendered_breakdown} if tendered_breakdown else {},
		'cash_change_breakdowns': [change_breakdown] if change_breakdown else [],
	}).content.decode('utf-8')
	return JsonResponse({'ok': True, 'sale_id': sale.id, 'paid': float(amount_paid), 'balance': float(sale.balance), 'change_amount': float(change_amount), 'receipt_html': receipt_html})


@login_required
@transaction.atomic
def add_payment(request, pk):
	sale = get_object_or_404(Sale, pk=pk)
	if request.method != "POST":
		return JsonResponse({"error": "Méthode non autorisée."}, status=405)
	try:
		amount = Decimal(request.POST.get("amount", "0") or "0")
	except Exception:
		amount = Decimal("0")
	method = request.POST.get("method") or Payment.CASH

	if amount <= 0:
		messages.error(request, "Montant invalide.")
	else:
		pmt = Payment.objects.create(sale=sale, amount=amount, method=method)
		# Mouvement caisse
		record_payment_movement(pmt)
		messages.success(request, "Paiement ajouté.")
	return redirect("sale_list")


# Catégories CRUD
@login_required
def category_list(request):
	categories = Category.objects.all()
	return render(request, "store/category_list.html", {"categories": categories})


@login_required
def stock_movements_list(request):
	product_id = request.GET.get("product")
	reason = request.GET.get("reason")
	qs = StockMovement.objects.select_related("product", "sale")
	if product_id:
		qs = qs.filter(product_id=product_id)
	if reason:
		qs = qs.filter(reason=reason)
	products = Product.objects.order_by("name")
	return render(request, "store/stock_movements.html", {
		"movements": qs[:500],
		"products": products,
		"reason": reason or "",
		"product_id": product_id or "",
		"reasons": StockMovement.REASONS,
	})


@admin_required
def category_create(request):
	if request.method == "POST":
		name = (request.POST.get("name") or "").strip()
		desc = (request.POST.get("description") or "").strip()
		if not name:
			messages.error(request, "Le nom est requis.")
		else:
			Category.objects.create(name=name, description=desc)
			messages.success(request, "Catégorie créée.")
			return redirect("category_list")
	return render(request, "store/category_form.html", {"title": "Nouvelle catégorie"})


@admin_required
def category_create_ajax(request):
	if request.method != "POST":
		return JsonResponse({"error": "Méthode non autorisée."}, status=405)
	name = (request.POST.get("name") or "").strip()
	desc = (request.POST.get("description") or "").strip()
	if not name:
		return JsonResponse({"error": "Le nom est requis."}, status=400)
	c = Category.objects.create(name=name, description=desc)
	return JsonResponse({"id": c.id, "name": c.name})


@admin_required
def category_delete_ajax(request, pk):
	if request.method != "POST":
		return JsonResponse({"error": "Méthode non autorisée."}, status=405)
	category = get_object_or_404(Category, pk=pk)
	# Délier produits puis supprimer
	Product.objects.filter(category=category).update(category=None)
	category.delete()
	return JsonResponse({"ok": True})


@admin_required
def category_update(request, pk):
	category = get_object_or_404(Category, pk=pk)
	if request.method == "POST":
		name = (request.POST.get("name") or "").strip()
		desc = (request.POST.get("description") or "").strip()
		if not name:
			messages.error(request, "Le nom est requis.")
		else:
			category.name = name
			category.description = desc
			category.save()
			messages.success(request, "Catégorie mise à jour.")
			return redirect("category_list")
	return render(request, "store/category_form.html", {"title": f"Modifier {category.name}", "category": category})


@admin_required
def category_delete(request, pk):
	category = get_object_or_404(Category, pk=pk)
	if request.method == "POST":
		# Délier les produits
		Product.objects.filter(category=category).update(category=None)
		category.delete()
		messages.success(request, "Catégorie supprimée.")
		return redirect("category_list")
	return render(request, "store/confirm_delete.html", {"object": category, "type": "catégorie", "cancel_url": reverse("category_list")})


# Dettes
@login_required
def debts_list(request):
	all_sales = Sale.objects.select_related("customer").prefetch_related('items', 'payments')
	sales = [s for s in all_sales if s.balance > 0]
	total_debt = sum((s.balance for s in sales), start=Decimal("0"))
	return render(request, "store/debts_list.html", {"sales": sales, "total_debt": total_debt})


@login_required
@transaction.atomic
def bulk_settle_customer(request, pk):
	"""Régler plusieurs dettes d'un client en une seule opération.

	GET (AJAX): renvoie un fragment HTML listant les ventes impayées et le total dû.
	POST: reçoit un montant et une méthode, alloue du plus ancien au plus récent.
	"""
	customer = get_object_or_404(Customer, pk=pk)
	if request.method == "POST":
		try:
			amount = Decimal(request.POST.get("amount", "0") or "0")
		except Exception:
			amount = Decimal("0")
		method = request.POST.get("method") or Payment.CASH

		if amount <= 0:
			messages.error(request, "Montant invalide.")
			return redirect("sale_list")

		unpaid_sales = [s for s in Sale.objects.filter(customer=customer).order_by("created_at").prefetch_related('items', 'payments') if s.balance > 0]
		remaining = amount
		for s in unpaid_sales:
			if remaining <= 0:
				break
			bal = s.balance
			pay = bal if bal <= remaining else remaining
			if pay > 0:
				pmt = Payment.objects.create(sale=s, amount=pay, method=method)
				record_payment_movement(pmt)
				remaining -= pay

		paid_total = amount - remaining
		# Réponse JSON si AJAX
		if request.headers.get('x-requested-with') == 'XMLHttpRequest':
			return JsonResponse({
				"ok": True,
				"paid_total": float(paid_total),
				"remaining_input": float(remaining),
			})
		# Sinon, messages + redirection
		if paid_total > 0:
			messages.success(request, f"Paiement de {paid_total} € réparti sur les dettes de {customer}.")
		else:
			messages.info(request, f"Aucune dette à régler pour {customer}.")
		return redirect("sale_list")

	# GET: renvoyer un fragment HTML (modale) avec le détail des dettes
	unpaid_sales = [s for s in Sale.objects.filter(customer=customer).order_by("created_at") if s.balance > 0]
	total_debt = sum((s.balance for s in unpaid_sales), start=Decimal("0"))
	# Si AJAX, renvoyer le fragment; sinon, rediriger vers l'historique
	if request.headers.get('x-requested-with') == 'XMLHttpRequest':
		return render(request, "store/_bulk_payment.html", {
			"customer": customer,
			"sales": unpaid_sales,
			"total_debt": total_debt,
		})
	return redirect("sale_list")


@login_required
def debt_customers_list(request):
	"""Retourne un fragment HTML listant les clients ayant des dettes (>0)."""
	from collections import defaultdict
	debt_by_customer = defaultdict(Decimal)
	for s in Sale.objects.select_related("customer").prefetch_related('items', 'payments'):
		if s.customer_id and s.balance > 0:
			debt_by_customer[s.customer] += s.balance
	entries = sorted(debt_by_customer.items(), key=lambda x: (str(x[0]).lower()))
	return render(request, "store/_debt_customer_list.html", {"entries": entries})


# Exports Excel
@login_required
def export_sales_xlsx(request):
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from django.http import HttpResponse

	wb = Workbook()
	ws = wb.active
	ws.title = "Ventes"
	ws.append(["ID", "Date", "Client", "Total", "Payé", "Solde"])
	for s in Sale.objects.select_related("customer").prefetch_related('items', 'payments').order_by("-created_at"):
		ws.append([s.id, s.created_at.replace(tzinfo=None), str(s.customer or ""), float(s.total), float(s.amount_paid), float(s.balance)])
	for i in range(1, 7):
		ws.column_dimensions[get_column_letter(i)].width = 18
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="ventes.xlsx"'
	wb.save(response)
	return response


@login_required
def export_debts_xlsx(request):
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from django.http import HttpResponse

	wb = Workbook()
	ws = wb.active
	ws.title = "Dettes"
	ws.append(["Vente ID", "Date", "Client", "Total", "Payé", "Solde"])
	for s in Sale.objects.select_related("customer").prefetch_related('items', 'payments').order_by("-created_at"):
		if s.balance > 0:
			ws.append([s.id, s.created_at.replace(tzinfo=None), str(s.customer or ""), float(s.total), float(s.amount_paid), float(s.balance)])
	for i in range(1, 7):
		ws.column_dimensions[get_column_letter(i)].width = 18
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="dettes.xlsx"'
	wb.save(response)
	return response


@login_required
def export_debtors_xlsx(request):
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from django.http import HttpResponse
	from collections import defaultdict

	debt_by_customer = defaultdict(Decimal)
	for s in Sale.objects.select_related("customer").prefetch_related('items', 'payments'):
		if s.customer_id and s.balance > 0:
			debt_by_customer[s.customer] += s.balance

	wb = Workbook()
	ws = wb.active
	ws.title = "Clients endettés"
	ws.append(["Nom", "Prénom", "Dette"])
	for customer, amount in sorted(debt_by_customer.items(), key=lambda x: (str(x[0]).lower())):
		ln = customer.last_name or ""
		fn = customer.first_name or ""
		ws.append([ln, fn, float(amount)])

	# Formater la colonne Dette en Euro
	max_row = ws.max_row
	for r in range(2, max_row + 1):
		cell = ws.cell(row=r, column=3)
		# Format compatible et explicite
		cell.number_format = '#,##0.00 "€"'
	for i in range(1, 4):
		ws.column_dimensions[get_column_letter(i)].width = 22
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="clients_endettes.xlsx"'
	wb.save(response)
	return response


@login_required
def export_products_xlsx(request):
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from django.http import HttpResponse

	wb = Workbook()
	ws = wb.active
	ws.title = "Catalogue"
	ws.append(["Nom", "Catégorie", "Prix", "Quantité"])
	for p in Product.objects.select_related("category").order_by("category__name", "name"):
		cat = str(p.category) if p.category_id else ""
		ws.append([p.name, cat, float(p.price), int(p.quantity)])
	# Largeurs colonnes
	widths = [28, 22, 12, 12]
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = w
	# Format prix
	for r in range(2, ws.max_row + 1):
		ws.cell(row=r, column=3).number_format = '#,##0.00 "€"'

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="catalogue_produits.xlsx"'
	wb.save(response)
	return response


# Ticket PDF
@login_required
def sale_receipt_pdf(request, pk):
	from io import BytesIO
	from reportlab.pdfgen import canvas
	from reportlab.lib.units import mm
	from reportlab.lib.utils import ImageReader
	from django.http import HttpResponse
	from django.conf import settings
	import os

	sale = get_object_or_404(Sale.objects.select_related("customer"), pk=pk)

	# Mise en page style "ticket 80mm"
	W = 80 * mm  # largeur ~ 80mm
	left = 6 * mm
	right = W - 6 * mm
	top_margin = 8 * mm
	bottom_margin = 8 * mm
	lh = 12  # line height

	def fmt(v: Decimal) -> str:
		return f"{v:.2f} €"

	# Estimer la hauteur du ticket avant de créer la page
	# Estimer précisément la hauteur utilisée (inclut le logo s'il existe)
	payments = list(sale.payments.all())
	logo_h_est = 0
	try:
		logo_path_est = os.path.join(settings.BASE_DIR, "static", "img", "logo.png")
		if os.path.exists(logo_path_est):
			from reportlab.lib.utils import ImageReader as _IR
			_reader = _IR(logo_path_est)
			iw, ih = _reader.getSize()
			desired_w = 28 * mm
			aspect = ih / float(iw or 1)
			logo_h_est = desired_w * aspect + 4  # 4 pts d'espacement
	except Exception:
		logo_h_est = 0

	used = 0
	used += logo_h_est
	# Titre + sous-titre
	used += lh  # COOP
	used += lh  # Reçu #
	# Date + Client
	used += lh * 2
	# Séparateur
	used += lh
	# Articles
	items_count = sale.items.count()
	used += items_count * lh
	# Séparateur
	used += lh
	# Totaux: Total, Payé, Dette
	used += 3 * lh
	# Paiements
	if payments:
		used += lh  # titre "Paiements"
		used += len(payments) * lh
	else:
		used += lh  # petite séparation
	# Footer (merci + ref) + petite marge
	used += 2 + 2 * lh

	H = top_margin + used + bottom_margin

	buffer = BytesIO()
	c = canvas.Canvas(buffer, pagesize=(W, H))

	# Fonts monospace pour look "ticket"
	c.setFont("Courier-Bold", 12)
	y = H - top_margin

	# Logo (si disponible)
	try:
		logo_path = os.path.join(settings.BASE_DIR, "static", "img", "logo.png")
		if os.path.exists(logo_path):
			reader = ImageReader(logo_path)
			iw, ih = reader.getSize()
			desired_w = 28 * mm
			aspect = ih / float(iw or 1)
			w = desired_w
			h = desired_w * aspect
			x = (W - w) / 2.0
			c.drawImage(reader, x, y - h, w, h, mask='auto', preserveAspectRatio=True, anchor='nw')
			y -= (h + 4)
	except Exception:
		pass

	# En-tête centré
	title = "COOP"
	tw = c.stringWidth(title, "Courier-Bold", 12)
	c.drawString((W - tw) / 2, y, title)
	y -= lh
	c.setFont("Courier", 10)
	subtitle = f"Reçu #{sale.id}"
	sw = c.stringWidth(subtitle, "Courier", 10)
	c.drawString((W - sw) / 2, y, subtitle)
	y -= lh

	# Infos
	c.drawString(left, y, f"Date : {sale.created_at:%d/%m/%Y %H:%M}")
	y -= lh
	customer_str = str(sale.customer or "") or "(aucun)"
	c.drawString(left, y, f"Client: {customer_str}")
	y -= lh

	# Séparateur
	c.drawString(left, y, "-" * 36)
	y -= lh

	# Articles
	c.setFont("Courier", 10)
	for it in sale.items.all():
		# Nom (tronqué si trop long)
		name = str(it.product)
		max_name = 22
		if len(name) > max_name:
			name = name[:max_name - 1] + "…"
		# gauche: "Nom xQ @P" ; droite: total
		left_text = f"{name} x{it.quantity} {it.unit_price:.2f}"
		right_text = f"{it.line_total:.2f}"
		c.drawString(left, y, left_text)
		rw = c.stringWidth(right_text, "Courier", 10)
		c.drawString(right - rw, y, right_text)
		y -= lh

	# Séparateur
	c.drawString(left, y, "-" * 36)
	y -= lh

	# Totaux
	c.setFont("Courier-Bold", 11)
	for label, value in (
		("Total", sale.total),
		("Payé", sale.amount_paid),
		("Dette", sale.balance),
	):
		c.drawString(left, y, f"{label}")
		val = f"{value:.2f} €"
		rw = c.stringWidth(val, "Courier-Bold", 11)
		c.drawString(right - rw, y, val)
		y -= lh

	# Paiements
	c.setFont("Courier", 10)
	if payments:
		c.drawString(left, y, "Paiements")
		y -= lh
		for pmt in payments:
			txt = f"{pmt.created_at:%d/%m %H:%M} {pmt.get_method_display():>6}"
			c.drawString(left, y, txt)
			amt = f"{pmt.amount:.2f} €"
			rw = c.stringWidth(amt, "Courier", 10)
			c.drawString(right - rw, y, amt)
			y -= lh
	else:
		# petite séparation visuelle
		c.drawString(left, y, "-" * 36)
		y -= lh

	# Footer
	y -= 2
	c.setFont("Courier", 9)
	msg = "Merci de votre achat !"
	mw = c.stringWidth(msg, "Courier", 9)
	c.drawString((W - mw) / 2, y, msg)
	y -= lh
	code = f"Ref: {sale.id}"
	cw = c.stringWidth(code, "Courier", 9)
	c.drawString((W - cw) / 2, y, code)

	c.showPage()
	c.save()
	pdf = buffer.getvalue()
	buffer.close()

	response = HttpResponse(content_type='application/pdf')
	response['Content-Disposition'] = f'inline; filename="ticket_vente_{sale.id}.pdf"'
	response.write(pdf)
	return response


# Tableau de bord
@login_required
def dashboard(request):
	from django.utils import timezone
	from datetime import timedelta, datetime, time
	now = timezone.now()
	start_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
	start_month = start_day.replace(day=1)
	start_year = start_day.replace(month=1, day=1)

	# Période sélectionnée (jour/semaine/mois/personnalisée)
	period = request.GET.get("p", "month")
	period_label = "Ce mois"
	custom_from = request.GET.get("from") or ""
	custom_to = request.GET.get("to") or ""
	if period == "day":
		start_period = start_day; end_period = now
		period_label = "Aujourd'hui"
	elif period == "week":
		start_period = start_day - timedelta(days=6); end_period = now
		period_label = "7 derniers jours"
	elif period == "month":
		start_period = start_month; end_period = now
		period_label = "Ce mois"
	elif period == "year":
		start_period = start_year; end_period = now
		period_label = "Cette année"
	elif period == "custom":
		try:
			from_date = datetime.strptime(custom_from, "%Y-%m-%d").date()
			exp_to = datetime.strptime(custom_to, "%Y-%m-%d").date()
		except Exception:
			# défaut: 7 derniers jours
			from_date = timezone.localdate() - timedelta(days=6)
			exp_to = timezone.localdate()
		# bornes inclusives/exclusives sur datetimes
		start_period = timezone.make_aware(datetime.combine(from_date, time.min))
		end_period = timezone.make_aware(datetime.combine(exp_to + timedelta(days=1), time.min))
		period_label = "Période personnalisée"
	else:
		start_period = start_month; end_period = now
		period_label = "Ce mois"

	from collections import defaultdict

	sales_today = list(
		Sale.objects.filter(created_at__gte=start_day)
		.prefetch_related('items', 'payments')
	)
	sum_today = sum((s.total for s in sales_today), start=Decimal("0"))
	count_today = len(sales_today)
	avg_today = (sum_today / count_today) if count_today else Decimal("0")

	sales_month = list(
		Sale.objects.filter(created_at__gte=start_month)
		.prefetch_related('items', 'payments')
	)
	sum_month = sum((s.total for s in sales_month), start=Decimal("0"))
	count_month = len(sales_month)
	avg_month = (sum_month / count_month) if count_month else Decimal("0")

	sales_period = list(
		Sale.objects.filter(created_at__gte=start_period, created_at__lt=end_period)
		.select_related('customer')
		.prefetch_related('items__product', 'payments')
	)
	sum_period = sum((s.total for s in sales_period), start=Decimal("0"))
	count_period = len(sales_period)
	avg_period = (sum_period / count_period) if count_period else Decimal("0")

	all_sales_debt = Sale.objects.prefetch_related('items', 'payments').select_related('customer')
	total_debt = Decimal("0")
	debt_by_customer = defaultdict(Decimal)
	for s in all_sales_debt:
		bal = s.balance
		if bal > 0:
			total_debt += bal
			if s.customer_id:
				debt_by_customer[s.customer] += bal

	debtors = sorted(debt_by_customer.items(), key=lambda x: x[1], reverse=True)[:10]
	debtors_count = len(debt_by_customer)

	# Ventes récentes dans la période sélectionnée
	recent_sales = (
		Sale.objects.select_related("customer")
		.filter(created_at__gte=start_period, created_at__lt=end_period)
		.order_by("-created_at")[:10]
	)

	# Graphique des ventes synchronisé avec la période
	chart_labels = []
	chart_values = []
	if period == "day":
		# 24 heures
		amount_per_hour = [Decimal("0")] * 24
		for s in sales_period:
			h = s.created_at.hour
			amount_per_hour[h] += s.total
		chart_labels = [f"{h:02d}h" for h in range(24)]
		chart_values = [float(x) for x in amount_per_hour]
	elif period == "year":
		# Regroupement par mois
		months = []  # (label, start, end)
		m_cursor = start_year.replace(day=1)
		while m_cursor < end_period:
			# début mois courant
			m_start = m_cursor
			# calcul début mois suivant
			if m_cursor.month == 12:
				next_month = m_cursor.replace(year=m_cursor.year + 1, month=1, day=1)
			else:
				next_month = m_cursor.replace(month=m_cursor.month + 1, day=1)
			m_end = next_month
			months.append((m_start.strftime('%b'), m_start, m_end))
			m_cursor = next_month
		chart_labels = [lbl for (lbl, _s, _e) in months]
		chart_values = []
		for (_lbl, m_start, m_end) in months:
			total_m = sum((s.total for s in sales_period if (s.created_at >= m_start and s.created_at < m_end)), start=Decimal("0"))
			chart_values.append(float(total_m))
	else:
		# Regrouper par jour
		# Construire la liste des jours inclus
		date_cursor = start_period.date()
		end_date = (end_period - timedelta(seconds=1)).date()
		amount_per_day = []
		while date_cursor <= end_date:
			day_start = timezone.make_aware(datetime.combine(date_cursor, time.min))
			day_end = day_start + timedelta(days=1)
			total_d = sum((s.total for s in sales_period if day_start <= s.created_at < day_end), start=Decimal("0"))
			amount_per_day.append((date_cursor.strftime('%d/%m'), float(total_d)))
			date_cursor += timedelta(days=1)
		chart_labels = [d for d, _ in amount_per_day]
		chart_values = [v for _, v in amount_per_day]

	# Top produits (quantités) et meilleurs clients (CA) sur la période
	qty_per_product = defaultdict(int)
	amount_per_customer = defaultdict(Decimal)
	active_customers = set()
	for s in sales_period:
		for it in s.items.all():  # items already prefetched, no N+1
			qty_per_product[str(it.product)] += it.quantity
		if s.customer_id:
			amount_per_customer[str(s.customer)] += s.total
			active_customers.add(s.customer_id)
	# top 5
	top_products = sorted(qty_per_product.items(), key=lambda x: x[1], reverse=True)[:5]
	top_products_labels = [k for k, _ in top_products]
	top_products_values = [v for _, v in top_products]
	top_customers = sorted(amount_per_customer.items(), key=lambda x: x[1], reverse=True)[:5]
	top_customers_labels = [k for k, _ in top_customers]
	top_customers_values = [float(v) for _, v in top_customers]
	top_customers_list = [(k, float(v)) for k, v in top_customers]
	active_customers_count = len(active_customers)

	# Seuils par produit
	low_stock = Product.objects.filter(quantity__lte=F('min_stock')).order_by("quantity", "name")[:10]
	if low_stock:
		messages.warning(request, f"{len(low_stock)} produit(s) sous le seuil d'alerte stock.")
	# Alertes: on conserve seulement les gros paniers (> 50€) sur la période
	old_debts = []
	big_baskets = [s for s in sales_period if s.total >= Decimal("50")]  # seuil simple, ajustable
	# Top 3 des plus gros paniers sur la période sélectionnée
	biggest_baskets = []
	if sales_period:
		biggest_baskets = sorted(sales_period, key=lambda s: s.total, reverse=True)[:3]

	context = {
		"sum_today": sum_today,
		"sum_month": sum_month,
		"total_debt": total_debt,
		"period": period,
		"period_label": period_label,
		"sum_period": sum_period,
		"count_period": count_period,
		"avg_period": avg_period,
		"custom_from": custom_from,
		"custom_to": custom_to,
		"count_today": count_today,
		"count_month": count_month,
		"avg_today": avg_today,
		"avg_month": avg_month,
		"debtors": debtors,
		"debtors_count": debtors_count,
		"recent_sales": recent_sales,
		"chart_sales_labels": chart_labels,
		"chart_sales_data": chart_values,
		"top_products_labels": top_products_labels,
		"top_products_data": top_products_values,
		"top_customers_labels": top_customers_labels,
		"top_customers_data": top_customers_values,
		"top_customers": top_customers_list,
		"active_customers_count": active_customers_count,
		"low_stock": low_stock,
		"old_debts": old_debts,
		"big_baskets": big_baskets,
		"biggest_baskets": biggest_baskets,
	}
	return render(request, "store/dashboard.html", context)


# ------------------ CAISSE (simplifiée sans ouverture/fermeture explicite) ------------------
@login_required
def cash_dashboard(request):
	session = get_active_cash_session()
	denom_entries = _build_denom_entries()
	if not session:
		return render(request, 'store/cash_dashboard.html', {
			'has_session': False,
			'denom_entries': denom_entries,
		})
	# Comptage courant (ce que l'on a maintenant)
	current = session.current_counts or {}
	cash_current = current.get('cash') or {}
	cheques_count_current = current.get('cheques') or 0
	# Total espèces courant
	cash_total_current = Decimal('0')
	for denom, qty in cash_current.items():
		try:
			cash_total_current += Decimal(str(denom)) * Decimal(str(qty))
		except Exception:
			continue
	# Total chèques = somme des paiements méthode CHEQUE sur la session
	cheque_total = session.movements_total_cheque
	global_total = cash_total_current + cheque_total
	recent_movements = session.movements.select_related('sale', 'payment')[:25]
	# Ventilation billets/pièces actuelle
	bills = []
	coins = []
	for denom, qty in cash_current.items():
		try:
			val = Decimal(str(denom))
		except Exception:
			continue
		val_str = str(denom).replace('.', '_')
		png_path = settings.BASE_DIR / 'static' / 'img' / 'denoms' / f'{val_str}.png'
		svg_path = settings.BASE_DIR / 'static' / 'img' / 'denoms' / f'{val_str}.svg'
		if png_path.exists():
			img_rel = f'img/denoms/{val_str}.png'
		elif svg_path.exists():
			img_rel = f'img/denoms/{val_str}.svg'
		else:
			img_rel = None
		entry = {'denom': denom, 'qty': qty, 'img': img_rel}
		if val >= Decimal('5'):
			bills.append(entry)
		else:
			coins.append(entry)
	return render(request, 'store/cash_dashboard.html', {
		'has_session': True,
		'session': session,
		'cash_current_map': cash_current,
		'bills_current': bills,
		'coins_current': coins,
		'cheques_count_current': cheques_count_current,
		'cash_total_current': cash_total_current,
		'cheque_total': cheque_total,
		'global_total': global_total,
		'recent_movements': recent_movements,
		'denom_entries': denom_entries,
	})


@login_required
def cash_reset(request):
	"""Réinitialise la caisse avec un nouveau fond (fond de caisse). Ferme l'ancienne session si existante."""
	if request.method == 'POST':
		# Fermer ancienne session sans exposer le concept à l'utilisateur
		old = get_active_cash_session()
		if old:
			old.is_closed = True
			old.closed_at = timezone.now()
			old.save(update_fields=['is_closed', 'closed_at'])
		# Construire opening_counts
		cash_map = {}
		for denom in DENOMINATIONS:
			field_name = f'denom_{str(denom).replace(".", "_")}'
			try:
				val = int(request.POST.get(field_name, '0') or '0')
			except Exception:
				val = 0
			if val > 0:
				cash_map[str(denom)] = val
		try:
			cheques_count = int(request.POST.get('cheques_count', '0') or '0')
		except Exception:
			cheques_count = 0
		opening_counts = {'cash': cash_map, 'cheques': cheques_count}
		CashSession.objects.create(opening_counts=opening_counts, opening_note=None)
		messages.success(request, 'Fond de caisse initialisé.')
		return redirect('cash_dashboard')

	return render(request, 'store/cash_reset.html', {
		'denom_entries': [{ 'value': d, 'safe': str(d).replace('.', '_') } for d in DENOMINATIONS],
	})


@login_required
def cash_drop(request):
	"""Remise de caisse (retrait d'espèces). On renseigne les quantités retirées; calcule montant et crée mouvement DROP."""
	session = get_active_cash_session()
	if not session:
		messages.error(request, 'Aucun fond de caisse. Initialisez la caisse d’abord.')
		return redirect('cash_dashboard')
	if request.method == 'POST':
		removed = {}
		amount = Decimal('0')
		for denom in DENOMINATIONS:
			fname = f'denom_{str(denom).replace(".", "_")}'
			try:
				qty = int(request.POST.get(fname, '0') or '0')
			except Exception:
				qty = 0
			if qty > 0:
				removed[str(denom)] = qty
				amount += denom * qty
		note = (request.POST.get('note') or '').strip() or 'Remise de caisse'
		CashMovement.objects.create(
			session=session,
			movement_type=CashMovement.DROP,
			method=Payment.CASH,
			amount=amount,
			counts_snapshot={'cash_removed': removed},
			note=note,
		)
		messages.success(request, f'Remise de caisse enregistrée ({amount} €).')
		return redirect('cash_dashboard')

	return render(request, 'store/cash_drop.html', {
		'denom_entries': [{ 'value': d, 'safe': str(d).replace('.', '_') } for d in DENOMINATIONS],
	})


@login_required
def export_cash_movements_xlsx(request):
	"""Export des mouvements de caisse de la session active en XLSX."""
	session = get_active_cash_session()
	if not session:
		messages.error(request, 'Pas de session de caisse active.')
		return redirect('cash_dashboard')
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter
	from django.http import HttpResponse
	wb = Workbook()
	ws = wb.active
	ws.title = 'Mouvements caisse'
	ws.append(['ID', 'Date', 'Type', 'Méthode', 'Montant', 'Note'])
	for m in session.movements.all().order_by('-created_at'):
		ws.append([m.id, m.created_at.replace(tzinfo=None), m.get_movement_type_display(), m.get_method_display() if m.method else '', float(m.amount), m.note or ''])
	for i in range(1, 7):
		ws.column_dimensions[get_column_letter(i)].width = 20
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="mouvements_caisse.xlsx"'
	wb.save(response)
	return response


# ---- Réglages utilisateurs ----

@admin_required
def settings_users(request):
    users = User.objects.prefetch_related('groups').order_by('username')
    groups = Group.objects.all()
    return render(request, 'store/settings_users.html', {'users': users, 'groups': groups})


@admin_required
@require_POST
@transaction.atomic
def settings_user_create(request):
    username = request.POST.get('username', '').strip()
    password = request.POST.get('password', '').strip()
    group_id = request.POST.get('group', '')
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    if not username or not password:
        messages.error(request, 'Nom d\'utilisateur et mot de passe requis.')
        return redirect('settings_users')
    if User.objects.filter(username=username).exists():
        messages.error(request, f'Le nom d\'utilisateur « {username} » est déjà utilisé.')
        return redirect('settings_users')
    user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)
    if group_id:
        try:
            user.groups.set([Group.objects.get(pk=group_id)])
        except Group.DoesNotExist:
            pass
    messages.success(request, f'Utilisateur « {username} » créé.')
    return redirect('settings_users')


@admin_required
@require_POST
@transaction.atomic
def settings_user_update(request, pk):
    user = get_object_or_404(User, pk=pk)
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    group_id = request.POST.get('group', '')
    password = request.POST.get('password', '').strip()
    user.first_name = first_name
    user.last_name = last_name
    if password:
        user.set_password(password)
    user.save()
    if group_id:
        try:
            user.groups.set([Group.objects.get(pk=group_id)])
        except Group.DoesNotExist:
            user.groups.clear()
    else:
        user.groups.clear()
    messages.success(request, f'Utilisateur « {user.username} » mis à jour.')
    return redirect('settings_users')


@admin_required
@require_POST
def settings_user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        messages.error(request, 'Vous ne pouvez pas supprimer votre propre compte.')
        return redirect('settings_users')
    username = user.username
    user.delete()
    messages.success(request, f'Utilisateur « {username} » supprimé.')
    return redirect('settings_users')


@admin_required
@transaction.atomic
@require_POST
def sale_delete(request, pk):
	sale = get_object_or_404(Sale, pk=pk)
	restore_stock = request.POST.get('restore_stock') == '1'
	reverse_cash = request.POST.get('reverse_cash') == '1'

	if restore_stock:
		for item in sale.items.select_related('product').all():
			Product.objects.filter(pk=item.product_id).update(quantity=F('quantity') + item.quantity)
			StockMovement.objects.create(
				product=item.product,
				change=item.quantity,
				reason=StockMovement.ADJUST,
				note=f'Annulation vente #{sale.id}',
			)

	if reverse_cash:
		session = get_active_cash_session()
		if session:
			for mvt in CashMovement.objects.filter(sale=sale):
				if mvt.movement_type == CashMovement.PAYMENT and mvt.tendered_breakdown:
					adjust_session_counts(session, removed=mvt.tendered_breakdown)
				elif mvt.movement_type == CashMovement.PAYMENT and mvt.method == 'CHEQUE':
					adjust_session_counts(session, cheques_delta=-1)
				elif mvt.movement_type == CashMovement.CHANGE and mvt.change_breakdown:
					adjust_session_counts(session, added=mvt.change_breakdown)
		CashMovement.objects.filter(sale=sale).delete()

	parts = []
	if restore_stock: parts.append('stock remis à jour')
	if reverse_cash: parts.append('caisse annulée')
	detail = ' (' + ', '.join(parts) + ')' if parts else ''
	sale.delete()
	messages.success(request, f'Vente #{pk} supprimée{detail}.')
	return redirect('sale_list')


@admin_required
@transaction.atomic
@require_POST
def product_adjust_stock(request, pk):
	product = get_object_or_404(Product, pk=pk)
	try:
		delta = int(request.POST.get('delta', '0') or '0')
	except ValueError:
		delta = 0
	note = (request.POST.get('note') or '').strip() or 'Ajustement manuel'
	if delta == 0:
		messages.error(request, 'Delta nul — aucun changement.')
		return redirect('product_list')
	new_qty = product.quantity + delta
	if new_qty < 0:
		messages.error(request, f'Stock insuffisant (actuel : {product.quantity}, delta : {delta}).')
		return redirect('product_list')
	Product.objects.filter(pk=product.pk).update(quantity=new_qty)
	StockMovement.objects.create(product=product, change=delta, reason=StockMovement.ADJUST, note=note)
	messages.success(request, f'Stock de « {product.name} » ajusté de {delta:+d} (nouveau stock : {new_qty}).')
	return redirect('product_list')

